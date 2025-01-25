from ..models import ValueStream
import pandas as pd
from openpyxl import load_workbook

queryset = ValueStream.objects.all()


def pragis_stavba_hours_worked(logic_name: str, file_path, attachment_date):
    first_day = attachment_date.replace(day=1)

    # Calculate the last day of the month
    if first_day.month == 12:  # Handle December separately
        # Increment the year, set to January, then subtract one day
        last_day = first_day.replace(year=first_day.year + 1, month=1, day=1) - pd.Timedelta(days=1)
    else:
        # Increment the month, then subtract one day
        last_day = first_day.replace(month=first_day.month + 1, day=1) - pd.Timedelta(days=1)

    if logic_name == "hours_worked":
        # Load the workbook and sheet
        wb = load_workbook(file_path)
        sheet = wb['List1']

        # Get the data from the range A4:AH8
        data = []
        for row in sheet['A4':'AH8']:
            data.append([cell.value for cell in row])
        # Create a DataFrame
        df = pd.DataFrame(data).T.fillna(0)

        date_range = pd.date_range(start=first_day, end=last_day, freq='D')
        # Create a DataFrame with the same number of rows as the date_range
        df_calendar = pd.DataFrame(index=date_range)
        print(f"First day is: {first_day}")
        print(f"Last day is: {last_day}")

        workable_columns_subset = df.rename(columns={0: 'Date', 2: 'Železný', 3: "Milko", 4: "Brinček"})[["Date", "Železný", "Milko", "Brinček"]]
        workable_columns_subset = workable_columns_subset.drop(index=0).reset_index(drop=True)
        # Assign the date_range to column 0
        df_calendar[0] = date_range
        workable_columns_subset = workable_columns_subset.drop(index=0).reset_index(drop=True)

        # Adjust rows in workable_columns_subset to match the date range
        if len(workable_columns_subset) > len(date_range):
            # Drop excess rows
            workable_columns_subset = workable_columns_subset.iloc[:len(date_range)]
        elif len(workable_columns_subset) < len(date_range):
            # Add empty rows to match date_range
            for _ in range(len(date_range) - len(workable_columns_subset)):
                workable_columns_subset = workable_columns_subset.append(
                    pd.Series([None] * len(workable_columns_subset.columns), index=workable_columns_subset.columns),
                    ignore_index=True)

        # Assign the date range to the 'Date' column
        workable_columns_subset.index = df_calendar.index
        workable_columns_subset["Date"] = date_range.strftime("%Y-%m-%d")

        # Set the 'Date' column as the index
        workable_columns_subset.set_index("Date", inplace=True)

        # print(workable_columns_subset)
        return workable_columns_subset

import xlrd

import xlrd
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
def pragis_voda_hours_worked(logic_name: str, file_path, attachment_date):
    print(f"Attachemnt date is: {attachment_date},\nIt is type of: {type(attachment_date)}")

    first_day = attachment_date.replace(day=1)

    # Calculate the last day of the month
    if first_day.month == 12:  # Handle December separately
        # Increment the year, set to January, then subtract one day
        last_day = first_day.replace(year=first_day.year + 1, month=1, day=1) - pd.Timedelta(days=1)
    else:
        # Increment the month, then subtract one day
        last_day = first_day.replace(month=first_day.month + 1, day=1) - pd.Timedelta(days=1)

    months = ["leden", "unor", "brezen", "duben",
              "kveten", "cerven", "cervenec", "zari",
              "rijen", "listopad", "prosinec"]

    if logic_name == "hours_worked":
        data = []

        if file_path.endswith('.xls'):
            # Open the .xls file using xlrd
            wb = xlrd.open_workbook(file_path)
            sheet = wb.sheet_by_name('List1')


            # Loop through the rows and columns manually
            for row_idx in range(4, 14):  # A5 corresponds to row 4, AG14 corresponds to row 13
                row_data = [sheet.cell_value(row_idx, col_idx) for col_idx in range(33)]  # Columns A to AG
                data.append(row_data)
        else:
            # Open the .xlsx file using openpyxl
            wb = load_workbook(file_path)
            sheet = wb['List1']

            # Get the data from the range A5:AG14
            for row in sheet['A5':'AG14']:
                data.append([cell.value for cell in row])

        # Create a DataFrame
        df = pd.DataFrame(data).T.fillna(0)

        date_range = pd.date_range(start=first_day, end=last_day, freq='D')
        # Create a DataFrame with the same number of rows as the date_range
        df_calendar = pd.DataFrame(index=date_range)
        print(f"First day is: {first_day}")
        print(f"Last day is: {last_day}")

        workable_columns_subset = df.rename(columns={0: 'Date', 2: 'Miro Bilý', 3: "Džurban", 4: "Pavol Miko", 5: "Štofko", 6: "Sinčák", 7: "Miko", 8: "Tekeli", 9: "Bilý Z."})[
            ["Date", "Miro Bilý", "Džurban", "Pavol Miko", "Štofko", "Sinčák", "Miko", "Tekeli", "Bilý Z."]]
        # Assign the date_range to column 0
        df_calendar[0] = date_range
        workable_columns_subset = workable_columns_subset.drop(index=0).reset_index(drop=True)


        # Adjust rows in workable_columns_subset to match the date range
        if len(workable_columns_subset) > len(date_range):
            # Drop excess rows
            workable_columns_subset = workable_columns_subset.iloc[:len(date_range)]
        elif len(workable_columns_subset) < len(date_range):
            # Add empty rows to match date_range
            for _ in range(len(date_range) - len(workable_columns_subset)):
                workable_columns_subset = workable_columns_subset.append(
                    pd.Series([None] * len(workable_columns_subset.columns), index=workable_columns_subset.columns),
                    ignore_index=True)

        # Assign the date range to the 'Date' column
        workable_columns_subset.index = df_calendar.index
        workable_columns_subset["Date"] = date_range.strftime("%Y-%m-%d")

        # Set the 'Date' column as the index
        workable_columns_subset.set_index("Date", inplace=True)

        #print(workable_columns_subset)
        return workable_columns_subset




"""for entry in queryset:
    subject_logic = entry.subject
    if subject_logic == "hours_worked":
        print(pragis_stavba_hours_worked(subject_logic,
                                   "/Users/robertsoroka/Downloads/září.xlsx"))
"""