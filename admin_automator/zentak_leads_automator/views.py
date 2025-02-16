from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from .playwright_service.playwright_nejstav_scraper_instance import NejstavScraper
from .models import Lead
from django.core import serializers
from datetime import datetime
#from .price_indication_service.price_indicator import process_lead, estimate_construction_activities
from django.views.generic import TemplateView, DetailView
from django.urls import reverse
import pandas as pd
import json
from .price_indication_service.price_indicator import ConstructionActivityEstimator
import os
import logging
import numpy as np
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
logger = logging.getLogger(__name__)

def scraper_form(request):
    """Render the scraper form page"""
    return render(request, 'scraper_form.html')


def run_scraper(request):
    if request.method == "POST":
        try:
            scraper = NejstavScraper()
            leads_data = scraper.scrape()
            
            # Create Lead objects from the data
            leads_created = []
            for lead_dict in leads_data:
                # Convert date from "DD.MM.YYYY HH:MM" to "YYYY-MM-DD"
                try:
                    date_str = lead_dict['post_date']
                    if date_str:
                        date_obj = datetime.strptime(date_str, "%d.%m.%Y %H:%M")
                        formatted_date = date_obj.strftime("%Y-%m-%d")
                    else:
                        formatted_date = None
                except ValueError:
                    print(f"Invalid date format: {date_str}")
                    formatted_date = None

                lead = Lead.objects.create(
                    headline=lead_dict['headline'],
                    description=lead_dict['description'],
                    post_date=formatted_date,  # Already formatted as YYYY-MM-DD
                    value=lead_dict['value'],
                    customer_name=lead_dict['customer_name'],
                    customer_email=lead_dict['customer_email'],
                    customer_phone=lead_dict['customer_phone'],
                    location=lead_dict['location']
                )
                leads_created.append(lead)
            
            # Convert leads to a list of dictionaries for JSON response
            leads_json = []
            for lead in leads_created:
                leads_json.append({
                    'headline': lead.headline,
                    'description': lead.description,
                    'post_date': lead.post_date,  # Already in correct format
                    'value': lead.value,
                    'customer_name': lead.customer_name,
                    'customer_email': lead.customer_email,
                    'customer_phone': lead.customer_phone,
                    'location': lead.location,
                    'created_at': lead.created_at.strftime('%Y-%m-%d %H:%M:%S')
                })
            
            return JsonResponse({
                "status": "Scraping completed!",
                "leads": leads_json
            })
            
        except Exception as e:
            import traceback
            return JsonResponse({
                "status": "Error",
                "message": str(e),
                "traceback": traceback.format_exc()
            }, status=500)
            
    return JsonResponse({"status": "Invalid request method. Use POST."})

from django.http import JsonResponse, HttpResponse
from django.views.generic import TemplateView, DetailView
from datetime import datetime
from openpyxl import Workbook
from .models import Lead

class LeadsListView(TemplateView):
    template_name = 'leads_list.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['leads'] = Lead.objects.all()
        return context

    def generate_excel(self):
        """Generate an Excel file containing the leads data using openpyxl directly."""
        try:
            # Get all leads
            leads = Lead.objects.all()
            
            if not leads.exists():
                return HttpResponse("No leads available to export.", status=204)

            # Create a new workbook and select the active sheet
            wb = Workbook()
            ws = wb.active
            ws.title = 'Leads'

            # Define headers
            headers = [
                'ID',
                'Headline',
                'Description',
                'Post Date',
                'Value',
                'Customer Name',
                'Customer Email',
                'Customer Phone',
                'Location',
                'Created At'
            ]

            # Write headers
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Write data
            for row, lead in enumerate(leads, 2):  # Start from row 2 (after headers)
                ws.cell(row=row, column=1, value=str(lead.lead_id))
                ws.cell(row=row, column=2, value=str(lead.headline))
                ws.cell(row=row, column=3, value=str(lead.description))
                ws.cell(row=row, column=4, value=lead.post_date.strftime('%Y-%m-%d') if lead.post_date else '')
                ws.cell(row=row, column=5, value=str(lead.value) if lead.value else '')
                ws.cell(row=row, column=6, value=str(lead.customer_name))
                ws.cell(row=row, column=7, value=str(lead.customer_email))
                ws.cell(row=row, column=8, value=str(lead.customer_phone))
                ws.cell(row=row, column=9, value=str(lead.location))
                ws.cell(row=row, column=10, value=lead.created_at.strftime('%Y-%m-%d %H:%M:%S') if lead.created_at else '')

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Create the response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=leads.xlsx'

            # Save the workbook to the response
            wb.save(response)

            return response

        except Exception as e:
            import traceback
            print(f"Error in generate_excel: {str(e)}")
            print(traceback.format_exc())
            return HttpResponse(
                f"Error creating Excel file: {str(e)}", 
                status=500
            )

    def post(self, request, *args, **kwargs):
        lead_id = request.POST.get('lead_id')
        if lead_id:
            try:
                return JsonResponse({"status": "Lead processed successfully."})
            except Exception as e:
                return JsonResponse({
                    "status": "Error",
                    "message": str(e)
                }, status=500)
        else:
            return JsonResponse({"status": "No lead ID provided."}, status=400)

    def get(self, request, *args, **kwargs):
        if 'export_excel' in request.GET:
            return self.generate_excel()
        return super().get(request, *args, **kwargs)

class LeadDetailView(DetailView):
    model = Lead
    template_name = 'lead_detail.html'  # Specify your template name here
    context_object_name = 'lead'  # This will be used in the template to refer to the lead object

    def get_object(self, queryset=None):
        # Override to get the lead by lead_id instead of the default id
        return Lead.objects.get(lead_id=self.kwargs['lead_id'])


class ConstructionProjectEstimateView(TemplateView):
    """
    View for estimating construction projects using LLM.
    Provides project analysis, cost estimation, and Excel export functionality.
    """
    template_name = 'construction_project_estimate.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        lead_id = self.request.GET.get('lead_id')
        if lead_id:
            try:
                lead = Lead.objects.get(lead_id=lead_id)
                context['lead'] = lead
            except Lead.DoesNotExist:
                context['error'] = f"Lead with ID {lead_id} not found"
        return context

    def estimate_project(self, project_inputs):
        try:
            # Get the absolute path to the context file
            current_dir = os.path.dirname(os.path.abspath(__file__))
            context_file = 'price_indication_service/construction_context.json'
            context_path = os.path.join(current_dir, context_file)

            logger.info(f"Looking for context file at: {context_path}")
            
            if not os.path.exists(context_path):
                raise FileNotFoundError(f"Context file not found at: {context_path}")
                
            estimator = ConstructionActivityEstimator(context_path)
            estimation = estimator.estimate_construction_activities(project_inputs)
            
            return estimation
            
        except Exception as e:
            logger.error(f"Estimation error: {str(e)}")
            raise

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            required_fields = ['description', 'value', 'location', 'headline']
            
            # Validate required fields
            missing_fields = [field for field in required_fields if not data.get(field)]
            if missing_fields:
                return JsonResponse({
                    "status": "Error",
                    "message": f"Missing required fields: {', '.join(missing_fields)}"
                }, status=400)

            # Convert and validate value
            try:
                data['value'] = float(data['value'])
            except (ValueError, TypeError):
                return JsonResponse({
                    "status": "Error",
                    "message": "Invalid value format - must be a number"
                }, status=400)

            try:
                estimation = self.estimate_project(data)
                if not estimation or not estimation.get('activities'):
                    return JsonResponse({
                        "status": "Error",
                        "message": "No activities generated in estimation"
                    }, status=400)
                
                excel_response = self.generate_excel(estimation)
                return excel_response
            
            except Exception as e:
                logger.error(f"Estimation processing error: {str(e)}", exc_info=True)
                return JsonResponse({
                    "status": "Error",
                    "message": f"Error processing estimation: {str(e)}"
                }, status=500)

        except json.JSONDecodeError:
            return JsonResponse({
                "status": "Error",
                "message": "Invalid JSON format"
            }, status=400)

    def generate_excel(self, estimation):
        """Generate Excel file containing detailed activity estimates."""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Výkaz výměr'
            
            # Define headers
            headers = [
                'Kategorie', 'Profese', 'Popis práce', 'Množství', 'MJ',
                'Jednotková cena', 'Celková cena', 'Materiál - nákup',
                'Práce - řemeslník', 'Navýšení - firma', 'Navýšení - materiál',
                'Jistota kalkulace (%)', 'Kategorie práce'
            ]
            
            # Write headers with formatting
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Write data
            activities = estimation.get('activities', [])
            total_price = 0  # Initialize total price counter
            
            for row_idx, activity in enumerate(activities, 2):
                # Calculate activity total price
                activity_price = float(str(activity.get('estimated_price', 0)).replace(',', '.'))
                total_price += activity_price  # Add to running total
                
                # Write activity data
                ws.cell(row=row_idx, column=1, value=str(activity.get('activity_category', '')))
                ws.cell(row=row_idx, column=2, value=str(activity.get('profession', '')))
                ws.cell(row=row_idx, column=3, value=str(activity.get('name', '')))
                ws.cell(row=row_idx, column=4, value=float(str(activity.get('quantity', 0)).replace(',', '.')))
                ws.cell(row=row_idx, column=5, value=str(activity.get('unit', '')))
                ws.cell(row=row_idx, column=6, value=float(str(activity.get('unit_price', 0)).replace(',', '.')))
                ws.cell(row=row_idx, column=7, value=activity_price)
                ws.cell(row=row_idx, column=8, value=float(str(activity.get('material_cost', 0)).replace(',', '.')))
                ws.cell(row=row_idx, column=9, value=float(str(activity.get('contractor_price', 0)).replace(',', '.')))
                ws.cell(row=row_idx, column=10, value=float(str(activity.get('company_markup', 0)).replace(',', '.')))
                ws.cell(row=row_idx, column=11, value=float(str(activity.get('material_markup', 0)).replace(',', '.')))
                ws.cell(row=row_idx, column=12, value=float(str(activity.get('confidence', 0)).replace(',', '.')))
                ws.cell(row=row_idx, column=13, value=str(activity.get('category', '')))

            # Format numbers
            for row in ws.iter_rows(min_row=2, max_row=len(activities)+1, min_col=4, max_col=11):
                for cell in row:
                    cell.number_format = '#,##0.00'
            
            # Add totals row with calculated total
            total_row = len(activities) + 2
            ws.cell(row=total_row, column=1, value='CELKEM').font = Font(bold=True)
            total_cell = ws.cell(row=total_row, column=7, value=total_price)
            total_cell.font = Font(bold=True)
            total_cell.number_format = '#,##0.00'
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Create metadata sheet with the same total price
            ws2 = wb.create_sheet(title='Metadata')
            metadata = {
                'Typ projektu': str(estimation.get('metadata', {}).get('project_type', '')),
                'Složitost': str(estimation.get('metadata', {}).get('complexity', '')),
                'Poznámky': str(estimation.get('metadata', {}).get('estimation_notes', '')),
                'Celková cena': total_price,  # Use the same total price
                'Celková jistota (%)': float(str(estimation.get('total_confidence', 0)).replace(',', '.'))
            }
            
            # Write metadata with formatting
            for row_idx, (key, value) in enumerate(metadata.items(), 1):
                ws2.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
                cell = ws2.cell(row=row_idx, column=2, value=value)
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=vykaz_vymer.xlsx'
            
            wb.save(response)
            return response

        except Exception as e:
            logger.error(f"Excel generation error: {str(e)}", exc_info=True)
            return JsonResponse({
                "status": "Error",
                "message": f"Error generating Excel file: {str(e)}"
            }, status=500)

    def get(self, request, *args, **kwargs):
        """
        Process GET request for viewing project estimate or downloading Excel file.
        """
        if 'export_excel' in request.GET:
            lead_id = request.GET.get('lead_id')
            if lead_id:
                try:
                    lead = Lead.objects.get(id=lead_id)
                    project_inputs = {
                        "description": lead.description,
                        "value": lead.value,
                        "location": lead.location,
                        "headline": lead.headline
                    }
                    estimation = self.estimate_project(project_inputs)
                except Lead.DoesNotExist:
                    return JsonResponse({
                        "status": "Error", 
                        "message": "Lead not found."
                    }, status=404)
            else:
                estimation = {
                    "activities": [],
                    "price_details": [],
                    "total_price": 0,
                    "total_confidence": 0,
                    "metadata": {}
                }
            return self.generate_excel(estimation)
        return super().get(request, *args, **kwargs)

    